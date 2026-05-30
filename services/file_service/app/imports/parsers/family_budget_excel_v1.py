from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass, field
from datetime import UTC, date, datetime, time
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

EXPECTED_HEADER = [
    "Дата операции",
    "Дата платежа",
    "Номер карты",
    "Статус",
    "Сумма операции",
    "Валюта операции",
    "Сумма платежа",
    "Валюта платежа",
    "Кэшбэк",
    "Категория",
    "MCC",
    "Описание",
    "Бонусы (включая кэшбэк)",
    "Округление на инвесткопилку",
    "Сумма операции с округлением",
]

SOURCE_TYPE = "excel_family_budget_v1"
SHEET_PATTERN = re.compile(r"^(Р|Д)_\d{2}\.\d{2}$")
MONEY_QUANT = Decimal("0.01")


@dataclass
class ValidationResult:
    ok: bool
    error_code: str | None = None
    message: str | None = None
    technical_details: str | None = None


@dataclass
class ParseIssue:
    sheet_name: str | None
    row_number: int | None
    column_name: str | None
    raw_value: str | None
    error_code: str
    message: str
    technical_details: str | None = None


@dataclass
class ParsedTransaction:
    source_sheet: str
    source_row_number: int
    type: str
    operation_at: datetime
    payment_at: datetime | None
    card_mask: str | None
    card_last4: str | None
    status: str | None
    operation_amount: Decimal
    operation_currency: str | None
    payment_amount: Decimal | None
    payment_currency: str | None
    cashback_amount: Decimal | None
    category_name: str | None
    mcc: str | None
    description: str | None
    bonus_amount: Decimal | None
    investment_rounding_amount: Decimal | None
    rounded_operation_amount: Decimal | None
    dedupe_key: str
    raw_payload: dict[str, str | None]

    def to_bulk_item(self, *, import_id: str, source_file_id: str) -> dict[str, Any]:
        return {
            "import_id": import_id,
            "source_file_id": source_file_id,
            "source_sheet": self.source_sheet,
            "source_row_number": self.source_row_number,
            "type": self.type,
            "operation_at": self.operation_at.isoformat(),
            "payment_at": self.payment_at.isoformat() if self.payment_at else None,
            "card_mask": self.card_mask,
            "card_last4": self.card_last4,
            "status": self.status,
            "operation_amount": _decimal_to_string(self.operation_amount),
            "operation_currency": self.operation_currency,
            "payment_amount": _decimal_to_string(self.payment_amount),
            "payment_currency": self.payment_currency,
            "cashback_amount": _decimal_to_string(self.cashback_amount),
            "category_name": self.category_name,
            "mcc": self.mcc,
            "description": self.description,
            "bonus_amount": _decimal_to_string(self.bonus_amount),
            "investment_rounding_amount": _decimal_to_string(self.investment_rounding_amount),
            "rounded_operation_amount": _decimal_to_string(self.rounded_operation_amount),
            "dedupe_key": self.dedupe_key,
            "raw_payload": self.raw_payload,
        }


@dataclass
class ParseResult:
    matched_sheets: list[str] = field(default_factory=list)
    total_rows: int = 0
    transactions: list[ParsedTransaction] = field(default_factory=list)
    errors: list[ParseIssue] = field(default_factory=list)

    @property
    def failed_rows(self) -> int:
        return sum(1 for error in self.errors if error.row_number is not None)


class FamilyBudgetExcelParser:
    source_type = SOURCE_TYPE

    def supports(self, file_metadata: dict, workbook_metadata: dict) -> bool:
        source_type = file_metadata.get("source_type")
        if source_type and source_type != SOURCE_TYPE:
            return False
        content_type = file_metadata.get("content_type")
        if content_type and content_type not in {
            SOURCE_TYPE,
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "application/vnd.ms-excel",
        }:
            return False
        sheet_names = workbook_metadata.get("sheet_names") or []
        return any(SHEET_PATTERN.match(name) for name in sheet_names)

    def validate_header(self, sheet_name: str, header_row: list[Any]) -> ValidationResult:
        header = _trim_trailing_empty(tuple(header_row))
        if list(header) != EXPECTED_HEADER:
            return ValidationResult(
                ok=False,
                error_code="unsupported_header",
                message=f"Лист {sheet_name} имеет неподдерживаемый набор колонок.",
                technical_details=f"Expected {EXPECTED_HEADER!r}, got {list(header)!r}",
            )
        return ValidationResult(ok=True)

    def inspect(self, file_bytes: bytes) -> ParseResult:
        return self.parse(file_bytes, rows=False)

    def parse(self, file_bytes: bytes, *, rows: bool = True) -> ParseResult:
        result = ParseResult()
        try:
            workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
        except Exception as exc:
            result.errors.append(
                ParseIssue(
                    sheet_name=None,
                    row_number=None,
                    column_name=None,
                    raw_value=None,
                    error_code="invalid_workbook",
                    message="Файл не удалось прочитать как Excel workbook.",
                    technical_details=str(exc),
                )
            )
            return result

        result.matched_sheets = [name for name in workbook.sheetnames if SHEET_PATTERN.match(name)]
        if not result.matched_sheets:
            result.errors.append(
                ParseIssue(
                    sheet_name=None,
                    row_number=None,
                    column_name=None,
                    raw_value=None,
                    error_code="no_supported_sheets",
                    message="В workbook не найдено листов формата Р_MM.YY или Д_MM.YY.",
                )
            )
            return result

        for sheet_name in result.matched_sheets:
            worksheet = workbook[sheet_name]
            iterator = worksheet.iter_rows(values_only=True)
            header = next(iterator, ())
            validation = self.validate_header(sheet_name, list(header))
            if not validation.ok:
                result.errors.append(
                    ParseIssue(
                        sheet_name=sheet_name,
                        row_number=1,
                        column_name=None,
                        raw_value=" | ".join(_stringify(value) or "" for value in header),
                        error_code=validation.error_code or "unsupported_header",
                        message=validation.message or f"Лист {sheet_name} имеет неподдерживаемый набор колонок.",
                        technical_details=validation.technical_details,
                    )
                )
                continue

            if not rows:
                continue

            for row_number, row in enumerate(iterator, start=2):
                if _is_empty_row(row):
                    continue
                result.total_rows += 1
                transaction, issues = self.normalize_row(sheet_name, row_number, row)
                if issues:
                    result.errors.extend(issues)
                elif transaction is not None:
                    result.transactions.append(transaction)
        return result

    def normalize_row(
        self,
        sheet_name: str,
        row_number: int,
        row: tuple[Any, ...],
    ) -> tuple[ParsedTransaction | None, list[ParseIssue]]:
        values = _row_values(row)
        raw_payload = {column: _stringify(values[column]) for column in EXPECTED_HEADER}
        issues: list[ParseIssue] = []

        operation_at = _parse_datetime(values["Дата операции"], "Дата операции", sheet_name, row_number, issues)
        payment_at = _parse_datetime(
            values["Дата платежа"],
            "Дата платежа",
            sheet_name,
            row_number,
            issues,
            required=False,
        )
        operation_amount = _parse_decimal(
            values["Сумма операции"],
            "Сумма операции",
            sheet_name,
            row_number,
            issues,
        )
        payment_amount = _parse_decimal(
            values["Сумма платежа"],
            "Сумма платежа",
            sheet_name,
            row_number,
            issues,
            required=False,
        )
        cashback_amount = _parse_decimal(
            values["Кэшбэк"],
            "Кэшбэк",
            sheet_name,
            row_number,
            issues,
            required=False,
        )
        bonus_amount = _parse_decimal(
            values["Бонусы (включая кэшбэк)"],
            "Бонусы (включая кэшбэк)",
            sheet_name,
            row_number,
            issues,
            required=False,
        )
        investment_rounding_amount = _parse_decimal(
            values["Округление на инвесткопилку"],
            "Округление на инвесткопилку",
            sheet_name,
            row_number,
            issues,
            required=False,
        )
        rounded_operation_amount = _parse_decimal(
            values["Сумма операции с округлением"],
            "Сумма операции с округлением",
            sheet_name,
            row_number,
            issues,
            required=False,
        )

        if issues or operation_at is None or operation_amount is None:
            return None, issues

        card_mask = _normalize_text(values["Номер карты"])
        payload_for_hash = "|".join(
            [
                sheet_name,
                str(row_number),
                operation_at.isoformat(),
                _decimal_to_string(operation_amount) or "",
                _normalize_text(values["Описание"]) or "",
            ]
        )
        return (
            ParsedTransaction(
                source_sheet=sheet_name,
                source_row_number=row_number,
                type="expense" if sheet_name.startswith("Р_") else "income",
                operation_at=operation_at,
                payment_at=payment_at,
                card_mask=card_mask,
                card_last4=_card_last4(card_mask),
                status=_normalize_text(values["Статус"]),
                operation_amount=operation_amount,
                operation_currency=_normalize_text(values["Валюта операции"]),
                payment_amount=payment_amount,
                payment_currency=_normalize_text(values["Валюта платежа"]),
                cashback_amount=cashback_amount,
                category_name=_normalize_text(values["Категория"]),
                mcc=_normalize_mcc(values["MCC"]),
                description=_normalize_text(values["Описание"]),
                bonus_amount=bonus_amount,
                investment_rounding_amount=investment_rounding_amount,
                rounded_operation_amount=rounded_operation_amount,
                dedupe_key=hashlib.sha256(payload_for_hash.encode("utf-8")).hexdigest(),
                raw_payload=raw_payload,
            ),
            [],
        )


def _row_values(row: tuple[Any, ...]) -> dict[str, Any]:
    padded = list(row[: len(EXPECTED_HEADER)])
    padded.extend([None] * (len(EXPECTED_HEADER) - len(padded)))
    return dict(zip(EXPECTED_HEADER, padded, strict=True))


def _is_empty_row(row: tuple[Any, ...]) -> bool:
    return all(value is None or value == "" for value in row[: len(EXPECTED_HEADER)])


def _trim_trailing_empty(values: tuple[Any, ...]) -> list[str | None]:
    trimmed = list(values)
    while trimmed and (trimmed[-1] is None or trimmed[-1] == ""):
        trimmed.pop()
    return [_stringify(value) for value in trimmed]


def _normalize_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _stringify(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return _as_utc(value).isoformat()
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=UTC).isoformat()
    text = str(value).strip()
    return text or None


def _as_utc(value: datetime) -> datetime:
    if value.tzinfo is None:
        return value.replace(tzinfo=UTC)
    return value.astimezone(UTC)


def _parse_datetime(
    value: Any,
    column: str,
    sheet_name: str,
    row_number: int,
    issues: list[ParseIssue],
    *,
    required: bool = True,
) -> datetime | None:
    if value in (None, ""):
        if required:
            issues.append(_issue(sheet_name, row_number, column, value, "missing_datetime"))
        return None
    if isinstance(value, datetime):
        return _as_utc(value)
    if isinstance(value, date):
        return datetime.combine(value, time.min, tzinfo=UTC)
    if isinstance(value, int | float):
        try:
            parsed = from_excel(value)
            return _as_utc(parsed)
        except Exception:
            pass
    text = str(value).strip()
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d.%m.%Y %H:%M:%S", "%d.%m.%Y"):
        try:
            parsed = datetime.strptime(text, fmt)
            return _as_utc(parsed)
        except ValueError:
            continue
    try:
        return _as_utc(datetime.fromisoformat(text))
    except ValueError:
        issues.append(_issue(sheet_name, row_number, column, value, "invalid_datetime"))
        return None


def _parse_decimal(
    value: Any,
    column: str,
    sheet_name: str,
    row_number: int,
    issues: list[ParseIssue],
    *,
    required: bool = True,
) -> Decimal | None:
    if value in (None, ""):
        if required:
            issues.append(_issue(sheet_name, row_number, column, value, "missing_decimal"))
        return None
    try:
        if isinstance(value, Decimal):
            parsed = value
        elif isinstance(value, int | float):
            parsed = Decimal(str(value))
        else:
            text = str(value).strip().replace("\xa0", "").replace(" ", "").replace(",", ".")
            parsed = Decimal(text)
        return parsed.quantize(MONEY_QUANT)
    except (InvalidOperation, ValueError) as exc:
        issues.append(_issue(sheet_name, row_number, column, value, "invalid_decimal", str(exc)))
        return None


def _normalize_mcc(value: Any) -> str | None:
    if value in (None, ""):
        return None
    try:
        decimal = Decimal(str(value).strip())
        if decimal == decimal.to_integral_value():
            return str(decimal.to_integral_value())
    except InvalidOperation:
        pass
    return str(value).strip()


def _card_last4(card_mask: str | None) -> str | None:
    if not card_mask:
        return None
    digits = re.sub(r"\D", "", card_mask)
    return digits[-4:] if len(digits) >= 4 else None


def _issue(
    sheet_name: str,
    row_number: int,
    column_name: str,
    raw_value: Any,
    error_code: str,
    technical_details: str | None = None,
) -> ParseIssue:
    return ParseIssue(
        sheet_name=sheet_name,
        row_number=row_number,
        column_name=column_name,
        raw_value=_stringify(raw_value),
        error_code=error_code,
        message=f"В строке {row_number} листа {sheet_name} значение в колонке '{column_name}' не удалось распознать.",
        technical_details=technical_details,
    )


def _decimal_to_string(value: Decimal | None) -> str | None:
    return None if value is None else format(value, "f")
